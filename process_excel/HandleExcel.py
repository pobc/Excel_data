"""
操作excel
1.读取excel内容
2.处理数据
3.存数据
"""
import openpyxl as xl

# xl.load_workbook(r"D:\work\text.xlsx").get_sheet_by_name("Sheet1")
# xl.load_workbook(r"D:\work\text.xlsx").get_sheet_by_name("Sheet2")

# 读取excel
sheets = xl.load_workbook(r"D:\work\text.xlsx")

# 定位sheet1
sheet1 = sheets.get_sheet_by_name("Sheet1")
"""
    print("abc 定位")
    print(sheet1["A2"].value)
    print(sheet1["B2"].value)
    print(sheet1["C2"].value)
    print(sheet1["D2"].value)
    print("====================")
    print(sheet1.cell(row=2, column=1).value)
    print(sheet1.cell(row=2, column=2).value)
    print(sheet1.cell(row=2, column=3).value)
    print(sheet1.cell(row=2, column=4).value)
"""
# 最大列
max_col_number = sheet1.max_column
# 最大行
max_row_number = sheet1.max_row
# ctrl + D                      约定
# print(max_col_number)
print("max_row_number:" + str(max_row_number))

# 最多读取多少列
max_read_count = 4

totalDataSave = []
for i in range(2, max_row_number + 1):
    rowData = []
    tableData = []
    for j in range(1, max_col_number + 1):
        rowValue = sheet1.cell(row=i, column=j).value
        # 只存前面4列的值
        if j <= max_read_count and rowValue is not None:
            rowData.append(rowValue)
            # 根据表的数量 进行复制  表名
        if j > max_read_count and rowValue is not None:
            tableData.append(rowValue)

    for vv in tableData:
        # 复制rowData 到 newRowData
        newRowData = rowData[:]
        newRowData.append(vv)
        totalDataSave.append(newRowData)
        print(newRowData)

    # 当没有表的时候，也要存在一行
    if tableData.__len__() == 0:
        newRowData = rowData[:]
        totalDataSave.append(newRowData)
        print(newRowData)

print("处理成功了，开始保存，将处理好的数据保存至新的excel")
# 新建excel
wb = xl.Workbook()
# 获取当前显示的sheet
newSheet = wb.active
# 将数据 保存进 sheet
title_list = ["一级域名", "二级域名", "三级域名", "四级域名", "表名"]
newSheet.append(title_list)
for hh in totalDataSave:
    newSheet.append(hh)
# 保存
wb.save(r"D:\work\ok_excel.xlsx")
print("保存成功")
